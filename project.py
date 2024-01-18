import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

import time

from openpyxl import Workbook, load_workbook 


service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)


#ATVER MĀJASLAPU
url = "https://www.binance.com/lv"
driver.get(url)
time.sleep(2)


#NORAIDA SĪKFAILUS
cookies=driver.find_element(By.ID, "onetrust-reject-all-handler")
cookies.click()
time.sleep(1)


#ATROD VALŪTU
crypto1=driver.find_elements(By.CLASS_NAME, "css-1ev4kiq")
crypto1[0].click()
time.sleep(1)


#ATROD CENU UN NOLASA TO
crypto1=driver.find_element(By.CLASS_NAME, "css-1bwgsh3")
valuta1=crypto1.text
print(valuta1)
time.sleep(1)


#AIZIET ATPAKAĻ
driver.get(url)
time.sleep(1)


crypto2=driver.find_elements(By.CLASS_NAME, "css-1ev4kiq")
crypto2[1].click()
time.sleep(1)


crypto2=driver.find_element(By.CLASS_NAME, "css-1bwgsh3")
valuta2=crypto2.text
print(valuta2)
time.sleep(1)


driver.get(url)
time.sleep(1)


crypto3=driver.find_elements(By.CLASS_NAME, "css-1ev4kiq")
crypto3[2].click()
time.sleep(1)


crypto3=driver.find_element(By.CLASS_NAME, "css-1bwgsh3")
valuta3=crypto3.text
print(valuta3)
time.sleep(1)


#ATVER EXCEL FAILU
project=load_workbook('project.xlsx')
ws=project.active
max_row=ws.max_row


#IEVIETO CENAS EXCEL FAILĀ
ws['A'+str(max_row+1)].value=valuta1
ws['B'+str(max_row+1)].value=valuta2
ws['C'+str(max_row+1)].value=valuta3


# KODS VĒRTĪBU IZMAIŅU APRĒĶINĀŠANAI  (NOŅEMT KOMENTĀRU, KAD PROGRAMMA IR IEGUVUSI DATUS)

#NOSAKA PIRMĀS VĒRTĪBAS
fvalue1=ws['A'+str(2)].value
fvalue2=ws['B'+str(2)].value
fvalue3=ws['C'+str(2)].value


#NOSAKA PĒDĒJĀS VĒRTĪBAS
lvalue1=ws['A'+str(max_row+1)].value
lvalue2=ws['B'+str(max_row+1)].value
lvalue3=ws['C'+str(max_row+1)].value


#PĀRVEIDO PIRMĀS VĒRTĪBAS UZ SKAITLI
fvf1=float(fvalue1.replace('$','').replace(',',''))
fvf2=float(fvalue2.replace('$','').replace(',',''))
fvf3=float(fvalue3.replace('$','').replace(',',''))


#PĀRVEIDO PĒDĒJĀS VĒRTĪBAS UZ SKAITLI
lvf1=float(lvalue1.replace('$','').replace(',',''))
lvf2=float(lvalue2.replace('$','').replace(',',''))
lvf3=float(lvalue3.replace('$','').replace(',',''))


#APRĒĶINA VĒRTĪBU IZMAIŅAS
valc1=(lvf1-fvf1)
valc2=(lvf2-fvf2)
valc3=(lvf2-fvf2)


#IEVIETO VĒRTĪBU IZMAIŅAS EXCEL FAILĀ
ws['D'+str(2)].value=valc1
ws['E'+str(2)].value=valc2
ws['F'+str(2)].value=valc3


#SAGLABĀ EXCEL FAILĀ "PROJECT" UN AIZVER PĀRLŪKPROGRAMMU
project.save('project.xlsx')
project.close()
driver.quit()


